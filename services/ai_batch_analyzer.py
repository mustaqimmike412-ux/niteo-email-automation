#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI 批量分析工具 v2.0
全面优化版本：
  1. 列名识别：支持模糊匹配、相似度计算、AI 辅助识别
  2. Excel 格式：支持多 sheet、合并单元格、多级表头、复杂布局
  3. AI 分类：更精准的行业识别、潜力评分、客户画像
"""

import os
import re
import json
import difflib
import pandas as pd
from typing import List, Dict, Tuple, Optional
from services.ai_extractor import AIExtractor
from services.file_content_extractor import extract_from_file
from utils.file_parser import parse_emails


class AIBatchAnalyzer:
    """AI 批量分析器 v2.0"""

    def __init__(self, api_key=None, base_url=None, model=None):
        self.extractor = AIExtractor(api_key=api_key, base_url=base_url, model=model)
        self.column_mapping = {}
        self.all_sheets_data = []  # 多 sheet 数据
        self.merged_cells_info = {}  # 合并单元格信息
        self.progress_callback = None  # 进度回调函数

    def is_available(self):
        return self.extractor.is_available()

    # ==================== 1. 智能列名识别（优化版）====================

    def analyze_columns(self, df: pd.DataFrame, use_ai: bool = True) -> Dict[str, str]:
        """
        智能分析 DataFrame 的列名，映射到标准字段
        支持：精确匹配、模糊匹配、相似度计算、AI 辅助识别
        """
        columns = [str(c).strip() for c in df.columns.tolist()]
        mapping = {}
        matched_cols = set()

        # 阶段 1: 精确正则匹配
        column_patterns = {
            'customer_name': [
                r'^客户$', r'^公司名称$', r'^Company$', r'^Customer$', r'^Name$',
                r'^客户名称$', r'^公司$', r'^Company\s*Name$', r'^客户公司$',
                r'^客户名$', r'^Company\s*Name\s*\(EN\)$', r'^客户（英文）$',
                r'^公司名$', r'^企业名称$', r'^企业$', r'^Company\s*Name\s*\(CN\)$',
                r'^公司（中文）$', r'^客户企业$'
            ],
            'website': [
                r'^网站$', r'^官网$', r'^Website$', r'^Web$', r'^URL$',
                r'^公司网站$', r'^Homepage$', r'^网址$', r'^Site$'
            ],
            'country': [
                r'^国家$', r'^地区$', r'^Country$', r'^Region$', r'^Nation$',
                r'^所在国家$', r'^国家/地区$', r'^国家地区$', r'^Country/Region$'
            ],
            'company_info': [
                r'^公司信息$', r'^描述$', r'^介绍$', r'^Company\s*Info$', r'^Description$',
                r'^公司简介$', r'^业务$', r'^Business$', r'^公司描述$', r'^产品信息$',
                r'^公司信息和客户主要产品$', r'^主要业务$', r'^经营范围$'
            ],
            'contact_info': [
                r'^联系人$', r'^联系人信息$', r'^Contact$', r'^Contacts$',
                r'^客户联系人$', r'^联系信息$', r'^Contact\s*Info$'
            ],
            'email': [
                r'^邮箱$', r'^邮件$', r'^Email$', r'^E-mail$', r'^电子邮件$',
                r'^客户邮箱$', r'^联系邮箱$', r'^Email\s*Address$', r'^邮箱地址$'
            ],
            'linkedin': [
                r'^领英$', r'^LinkedIn$', r'^Linkedin$', r'^领英邮箱$',
                r'^LinkedIn邮箱$', r'^LinkedIn\s*Email$'
            ],
            'phone': [
                r'^电话$', r'^手机$', r'^Phone$', r'^Tel$', r'^Contact\s*Number$',
                r'^联系方式$', r'^联系电话$', r'^Phone\s*Number$', r'^Mobile$'
            ],
            'address': [
                r'^地址$', r'^Address$', r'^Location$', r'^所在地$',
                r'^公司地址$', r'^办公地址$', r'^Office\s*Address$'
            ],
            'supplier': [
                r'^供应商$', r'^Supplier$', r'^供应商信息$', r'^Supplier\s*Info$'
            ],
            'customs_data': [
                r'^海关数据$', r'^Customs$', r'^海关$', r'^海关数据购买产品名$',
                r'^进口产品$', r'^Import\s*Products$'
            ],
            'logistics': [
                r'^物流信息$', r'^Logistics$', r'^物流$', r'^Shipping$',
                r'^物流方式$', r'^运输方式$'
            ]
        }

        for standard_name, patterns in column_patterns.items():
            for col in columns:
                if col in matched_cols:
                    continue
                for pattern in patterns:
                    if re.search(pattern, col, re.IGNORECASE):
                        mapping[standard_name] = col
                        matched_cols.add(col)
                        break
                if standard_name in mapping:
                    break

        # 阶段 2: 模糊匹配（相似度 >= 0.6）
        fuzzy_keywords = {
            'customer_name': ['客户', 'company', 'customer', 'name', '公司', '名称'],
            'website': ['website', 'web', 'url', 'site', '网站', '官网', '网址'],
            'country': ['country', 'nation', 'region', '国家', '地区', '所在'],
            'company_info': ['info', 'description', 'business', 'profile', '信息', '描述', '业务', '产品'],
            'email': ['email', 'mail', 'e-mail', '邮箱', '邮件', '电子邮'],
            'contact_info': ['contact', 'person', '联系人', '联系'],
            'phone': ['phone', 'tel', 'mobile', '电话', '手机', '联系'],
            'address': ['address', 'location', '地址', '所在地'],
            'linkedin': ['linkedin', '领英'],
            'supplier': ['supplier', '供应'],
            'customs_data': ['customs', '海关', '进口'],
            'logistics': ['logistics', '物流', '运输', 'shipping']
        }

        for standard_name, keywords in fuzzy_keywords.items():
            if standard_name in mapping:
                continue
            best_match = None
            best_score = 0
            for col in columns:
                if col in matched_cols:
                    continue
                col_lower = col.lower()
                for keyword in keywords:
                    # 计算相似度
                    score = difflib.SequenceMatcher(None, col_lower, keyword.lower()).ratio()
                    if score > best_score and score >= 0.6:
                        best_score = score
                        best_match = col
            if best_match:
                mapping[standard_name] = best_match
                matched_cols.add(best_match)

        # 阶段 3: AI 辅助识别（如果启用且未识别完整）
        if use_ai and self.is_available() and len(mapping) < 5:
            ai_mapping = self._ai_recognize_columns(columns, matched_cols)
            for key, val in ai_mapping.items():
                if key not in mapping and val not in matched_cols:
                    mapping[key] = val
                    matched_cols.add(val)

        self.column_mapping = mapping
        return mapping

    def _ai_recognize_columns(self, columns: List[str], matched_cols: set) -> Dict[str, str]:
        """使用 AI 辅助识别未匹配的列名"""
        unmatched = [c for c in columns if c not in matched_cols]
        if not unmatched:
            return {}

        system_prompt = """You are a data analyst. Given a list of column names from an Excel file, identify which ones correspond to standard fields.

Standard fields:
- customer_name: company/customer name
- website: company website/URL
- country: country/region
- company_info: company description/business info
- contact_info: contact person information
- email: email address
- linkedin: LinkedIn/email
- phone: phone number
- address: company address
- supplier: supplier information
- customs_data: customs/import data
- logistics: logistics/shipping info

Output ONLY valid JSON in this exact format:
{"customer_name": "original_column_name", "email": "original_column_name", ...}

Only include fields you are confident about. Use empty string "" if unsure."""

        user_prompt = f"""Column names: {json.dumps(unmatched, ensure_ascii=False)}

Map these to standard fields. Output JSON only."""

        try:
            content, error = self.extractor.llm._call(
                system_prompt, user_prompt,
                max_tokens=500, temperature=0.1, label='recognize_columns'
            )
            if error or not content:
                return {}

            # 清理 JSON
            if '```json' in content:
                content = content.split('```json')[1].split('```')[0]
            elif '```' in content:
                content = content.split('```')[1].split('```')[0]

            result = json.loads(content.strip())
            if isinstance(result, dict):
                # 过滤无效结果
                return {k: v for k, v in result.items()
                        if v and v in unmatched and k in [
                            'customer_name', 'website', 'country', 'company_info',
                            'contact_info', 'email', 'linkedin', 'phone',
                            'address', 'supplier', 'customs_data', 'logistics'
                        ]}
        except Exception:
            pass

        return {}

    # ==================== 2. 复杂 Excel 格式支持 ====================

    def analyze_file(self, file_path: str, sheet_name=None, header_row=None) -> List[Dict]:
        """
        分析文件，支持复杂格式

        Args:
            file_path: Excel 或 CSV 文件路径
            sheet_name: 指定 sheet 名称（None 则自动选择）
            header_row: 表头行号（None 则自动检测）

        Returns:
            分析结果列表
        """
        ext = os.path.splitext(file_path)[1].lower()

        if ext in ('.xlsx', '.xls'):
            return self._analyze_excel(file_path, sheet_name, header_row)
        elif ext == '.csv':
            return self._analyze_csv(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    def _analyze_excel(self, file_path: str, sheet_name=None, header_row=None) -> List[Dict]:
        """分析 Excel 文件，支持多 sheet、合并单元格、多级表头"""
        xl = pd.ExcelFile(file_path)
        all_results = []

        # 获取所有 sheet
        sheets = xl.sheet_names
        print(f"Excel 文件包含 {len(sheets)} 个 sheet: {sheets}")

        # 如果指定了 sheet_name
        if sheet_name and sheet_name in sheets:
            sheets_to_process = [sheet_name]
        else:
            # 自动选择数据量最大的 sheet
            sheets_to_process = self._select_best_sheets(xl, sheets)

        for sheet in sheets_to_process:
            try:
                print(f"\n处理 Sheet: {sheet}")

                # 检测表头行
                if header_row is None:
                    header_row = self._detect_header_row(file_path, sheet)
                    print(f"  自动检测到表头在第 {header_row + 1} 行")

                # 读取数据，处理合并单元格
                df = self._read_excel_with_merged_cells(file_path, sheet, header_row)

                if df.empty or len(df) < 1:
                    print(f"  Sheet {sheet} 为空，跳过")
                    continue

                # 处理多级表头
                df = self._flatten_multiheader(df)

                print(f"  读取到 {len(df)} 行, {len(df.columns)} 列")
                print(f"  列名: {df.columns.tolist()}")

                # 分析此 sheet（传递进度回调）
                sheet_results = self._analyze_dataframe(df, sheet_name=sheet, progress_callback=self.progress_callback)
                all_results.extend(sheet_results)

            except Exception as e:
                print(f"  处理 Sheet {sheet} 时出错: {e}")
                continue

        return all_results

    def _select_best_sheets(self, xl: pd.ExcelFile, sheets: List[str]) -> List[str]:
        """选择数据量最大的 sheet"""
        sheet_sizes = []
        for sheet in sheets:
            try:
                df = pd.read_excel(xl, sheet_name=sheet, nrows=5)
                # 估算总行数（通过文件大小或读取全部）
                df_full = pd.read_excel(xl, sheet_name=sheet)
                size = len(df_full) * len(df_full.columns)
                sheet_sizes.append((sheet, size, len(df_full)))
            except Exception:
                sheet_sizes.append((sheet, 0, 0))

        # 按数据量排序，选择最大的
        sheet_sizes.sort(key=lambda x: x[1], reverse=True)

        # 如果只有一个有意义的 sheet，返回它
        meaningful = [s for s in sheet_sizes if s[2] > 0]
        if len(meaningful) == 1:
            return [meaningful[0][0]]

        # 如果有多个，返回前 3 个（避免太多）
        return [s[0] for s in meaningful[:3]]

    def _detect_header_row(self, file_path: str, sheet: str) -> int:
        """自动检测表头行"""
        # 读取前 10 行尝试检测
        df_preview = pd.read_excel(file_path, sheet_name=sheet, nrows=10, header=None)

        best_row = 0
        best_score = 0

        for i in range(min(5, len(df_preview))):
            row = df_preview.iloc[i]
            # 计算这一行作为表头的得分
            score = 0

            for val in row:
                val_str = str(val).strip().lower() if pd.notna(val) else ''
                # 如果包含常见列名关键词，加分
                keywords = ['客户', '公司', 'email', '邮箱', 'name', '电话', '地址', 'country', 'website']
                for kw in keywords:
                    if kw in val_str:
                        score += 1

                # 如果值看起来像数据（包含 @ 或数字），减分
                if '@' in val_str or re.match(r'^\d', val_str):
                    score -= 2

            if score > best_score:
                best_score = score
                best_row = i

        return best_row

    def _read_excel_with_merged_cells(self, file_path: str, sheet: str, header_row: int) -> pd.DataFrame:
        """读取 Excel，处理合并单元格（将合并单元格的值填充到所有子单元格）"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            # 没有 openpyxl，使用 pandas 默认读取
            return pd.read_excel(file_path, sheet_name=sheet, header=header_row, dtype=str)

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet]

            # 读取合并单元格信息
            merged_ranges = ws.merged_cells.ranges
            merged_values = {}

            for merged_range in merged_ranges:
                # 获取合并区域的值（左上角单元格）
                min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                top_left_value = ws.cell(row=min_row, column=min_col).value

                # 记录合并区域的所有位置
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_values[(row, col)] = top_left_value

            # 读取所有数据
            data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
                row_data = []
                for col_idx, cell in enumerate(row, start=1):
                    # 优先使用合并单元格的值
                    if (row_idx, col_idx) in merged_values:
                        row_data.append(merged_values[(row_idx, col_idx)])
                    else:
                        row_data.append(cell.value)
                data.append(row_data)

            # 读取表头
            header = []
            for col_idx in range(1, len(data[0]) + 1 if data else 1):
                if (header_row + 1, col_idx) in merged_values:
                    header.append(merged_values[(header_row + 1, col_idx)])
                else:
                    val = ws.cell(row=header_row + 1, column=col_idx).value
                    header.append(val if val is not None else f'Column_{col_idx}')

            df = pd.DataFrame(data[1:], columns=header)  # 跳过表头行
            return df.astype(str)

        except Exception as e:
            print(f"  处理合并单元格时出错，使用默认读取: {e}")
            return pd.read_excel(file_path, sheet_name=sheet, header=header_row, dtype=str)

    def _flatten_multiheader(self, df: pd.DataFrame) -> pd.DataFrame:
        """展平多级表头"""
        # 检查是否有 MultiIndex 列
        if isinstance(df.columns, pd.MultiIndex):
            # 将多级表头合并为单级
            new_columns = []
            for col in df.columns:
                # 合并各级，去除空值和 NaN
                parts = [str(c).strip() for c in col if pd.notna(c) and str(c).strip()]
                new_col = ' '.join(parts) if parts else 'Unknown'
                new_columns.append(new_col)
            df.columns = new_columns

        # 清理列名
        df.columns = [str(c).strip() if pd.notna(c) else f'Column_{i}' for i, c in enumerate(df.columns)]

        return df

    def _analyze_csv(self, file_path: str) -> List[Dict]:
        """分析 CSV 文件"""
        encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312']

        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, dtype=str, encoding=encoding)
                if not df.empty:
                    print(f"CSV 编码: {encoding}")
                    return self._analyze_dataframe(df)
            except UnicodeDecodeError:
                continue

        raise ValueError("无法识别 CSV 文件编码")

    def _analyze_dataframe(self, df: pd.DataFrame, sheet_name: str = '', progress_callback=None) -> List[Dict]:
        """分析 DataFrame 数据（优化版：支持缓存和批量分类）"""
        if df.empty:
            return []

        # 智能识别列名
        column_mapping = self.analyze_columns(df)
        print(f"列名映射: {column_mapping}")

        # 初始化缓存
        try:
            from utils.import_cache import get_cache
            cache = get_cache()
        except ImportError:
            cache = None

        # 第一步：提取所有行的原始数据
        raw_results = []
        for idx, row in df.iterrows():
            try:
                # 跳过空行
                if row.isna().all() or all(str(v).strip() in ('', 'nan', 'None') for v in row if pd.notna(v)):
                    continue

                customer_data = self._extract_customer_from_row(row, column_mapping)
                contacts = self._extract_contacts_from_row(row, column_mapping)

                # 如果没有识别到客户名称，尝试从其他字段推断
                if not customer_data.get('customer_name'):
                    customer_data['customer_name'] = self._infer_customer_name(row, column_mapping)

                # 如果仍然没有客户名称，跳过
                if not customer_data.get('customer_name'):
                    continue

                raw_results.append({
                    'customer': customer_data,
                    'contacts': contacts,
                    'row_index': idx,
                    'sheet_name': sheet_name
                })

            except Exception as e:
                print(f"处理第 {idx + 1} 行时出错: {e}")
                continue

        # 第二步：使用缓存和批量分类优化AI调用
        if self.is_available() and raw_results:
            # 先检查缓存
            uncached_indices = []
            uncached_customers = []

            for i, item in enumerate(raw_results):
                customer = item['customer']
                cached = None
                if cache:
                    cached = cache.get(
                        customer['customer_name'],
                        customer.get('country', ''),
                        customer.get('company_info', '')
                    )

                if cached:
                    raw_results[i]['classification'] = cached
                    raw_results[i]['cache_hit'] = True
                else:
                    raw_results[i]['classification'] = None
                    raw_results[i]['cache_hit'] = False
                    uncached_indices.append(i)
                    uncached_customers.append({
                        'customer_name': customer['customer_name'],
                        'country': customer.get('country', ''),
                        'company_info': customer.get('company_info', ''),
                        'website': customer.get('website', ''),
                        'emails': item['contacts']
                    })

            # 批量分类未缓存的客户（每批10个，平衡速度和可靠性）
            BATCH_SIZE = 10
            total_batches = (len(uncached_customers) + BATCH_SIZE - 1) // BATCH_SIZE

            for batch_idx in range(total_batches):
                start = batch_idx * BATCH_SIZE
                end = min(start + BATCH_SIZE, len(uncached_customers))
                batch = uncached_customers[start:end]

                if progress_callback:
                    progress_callback(
                        'ai_classify',
                        f'AI智能分类中... (批次 {batch_idx + 1}/{total_batches}, {len(uncached_customers)}个客户)'
                    )

                try:
                    classifications = self.extractor.classify_customers_batch(batch)

                    # 将分类结果写回并缓存
                    for j, classification in enumerate(classifications):
                        result_idx = uncached_indices[start + j]
                        raw_results[result_idx]['classification'] = classification

                        # 写入缓存
                        if cache:
                            customer = raw_results[result_idx]['customer']
                            cache.set(
                                customer['customer_name'],
                                customer.get('country', ''),
                                customer.get('company_info', ''),
                                classification
                            )

                except Exception as e:
                    print(f"批量分类批次 {batch_idx + 1} 失败: {e}")
                    # 失败时使用默认分类
                    for j in range(len(batch)):
                        result_idx = uncached_indices[start + j]
                        raw_results[result_idx]['classification'] = self.extractor._default_classification()

        return raw_results

    def _infer_customer_name(self, row: pd.Series, column_mapping: Dict) -> str:
        """从行数据中推断客户名称（增强版）"""
        # 1. 尝试第一列
        if len(row) > 0 and pd.notna(row.iloc[0]):
            val = str(row.iloc[0]).strip()
            if val and val.lower() not in ('nan', 'none', 'null', ''):
                # 如果看起来像公司名称（包含 Corp, Inc, Ltd 等）
                if re.search(r'\b(Corp|Inc|Ltd|LLC|Co\.?|Company|Group)\b', val, re.IGNORECASE):
                    return val
                # 如果不太像邮箱或数字
                if '@' not in val and not val.isdigit():
                    return val

        # 2. 从网站域名推断公司名称
        website = None
        if 'website' in column_mapping:
            website_val = row.get(column_mapping['website'])
            if pd.notna(website_val):
                website = str(website_val).strip()
        
        if website:
            inferred_name = self._infer_name_from_website(website)
            if inferred_name:
                return inferred_name

        # 3. 扫描所有列，找最短且不像邮箱/数字的文本
        best_candidate = ''
        for val in row.values:
            if pd.notna(val):
                val_str = str(val).strip()
                if (val_str and val_str.lower() not in ('nan', 'none', 'null', '')
                        and '@' not in val_str and not val_str.isdigit()
                        and len(val_str) < len(best_candidate) or not best_candidate):
                    best_candidate = val_str

        return best_candidate

    def _infer_name_from_website(self, website: str) -> str:
        """从网站域名推断公司名称"""
        import re
        from urllib.parse import urlparse
        
        if not website:
            return ''
        
        try:
            # 解析 URL
            if not website.startswith(('http://', 'https://')):
                website = 'https://' + website
            
            parsed = urlparse(website)
            domain = parsed.netloc or parsed.path
            
            # 移除 www. 前缀
            domain = re.sub(r'^www\.', '', domain)
            
            # 获取主域名（移除 .com, .co.uk 等）
            parts = domain.split('.')
            if len(parts) >= 2:
                main_name = parts[0]
                
                # 将域名转换为可读的公司名称
                # 例如: solarpanda -> SOLAR PANDA, dlight -> D.LIGHT
                name = main_name.replace('-', ' ').replace('_', ' ')
                
                # 尝试将连续的单词分开（如 solarpanda -> solar panda）
                # 使用简单的启发式：在常见单词边界处分割
                import re
                # 尝试匹配常见模式：solar + panda, green + energy 等
                separated = re.sub(r'([a-z])([A-Z])', r'\1 \2', name)  # camelCase
                separated = re.sub(r'([a-zA-Z])([0-9])', r'\1 \2', separated)  # 字母+数字
                
                # 如果分离后的名称看起来合理，使用它
                if len(separated) > 2:
                    # 将每个单词首字母大写
                    words = separated.split()
                    capitalized = ' '.join(w.capitalize() for w in words)
                    return capitalized
                
                return name.capitalize()
        except Exception:
            pass
        
        return ''

    # ==================== 3. 数据提取（保持原有逻辑）====================

    def _extract_customer_from_row(self, row: pd.Series, column_mapping: Dict) -> Dict:
        """从一行数据中提取客户信息"""
        customer = {}

        fields = [
            ('customer_name', 'customer_name'),
            ('website', 'website'),
            ('country', 'country'),
            ('company_info', 'company_info'),
            ('address', 'address'),
            ('supplier', 'supplier'),
            ('customs_data', 'customs_data'),
            ('logistics', 'logistics')
        ]

        for standard_field, mapping_key in fields:
            if mapping_key in column_mapping:
                val = row.get(column_mapping[mapping_key])
                if pd.notna(val):
                    customer[standard_field] = str(val).strip()

        # 如果没有 customer_name，尝试第一列
        if 'customer_name' not in customer or not customer['customer_name']:
            if len(row) > 0 and pd.notna(row.iloc[0]):
                customer['customer_name'] = str(row.iloc[0]).strip()

        return customer

    def _extract_contacts_from_row(self, row: pd.Series, column_mapping: Dict) -> List[Dict]:
        """从一行数据中提取联系人信息（增强版）"""
        contacts = []
        sources = [
            ('contact_info', 'contact_info'),
            ('email', 'email_column'),
            ('linkedin', 'linkedin'),
            ('phone', 'phone')
        ]

        for mapping_key, source_name in sources:
            if mapping_key in column_mapping:
                val = row.get(column_mapping[mapping_key])
                if pd.notna(val):
                    text = str(val)
                    # 先尝试标准解析（含姓名/职位信息）
                    parsed = parse_emails(text)
                    if parsed:
                        for email, email_type, contact_name, job_title in parsed:
                            contacts.append({
                                'contact_name': contact_name or '',
                                'job_title': job_title or '',
                                'email_address': email,
                                'email_type': email_type,
                                'source': source_name
                            })
                    else:
                        # 如果标准解析失败，尝试纯邮箱列表解析
                        from utils.file_parser import parse_email_list
                        parsed_list = parse_email_list(text)
                        for email, email_type, contact_name, job_title in parsed_list:
                            contacts.append({
                                'contact_name': contact_name or '',
                                'job_title': job_title or '',
                                'email_address': email,
                                'email_type': email_type,
                                'source': source_name
                            })

        # 如果没有专门的列，扫描整行
        if not contacts:
            for col in row.index:
                val = row.get(col)
                if pd.notna(val):
                    text = str(val)
                    if '@' in text:
                        # 先尝试标准解析
                        parsed = parse_emails(text)
                        if parsed:
                            for email, email_type, contact_name, job_title in parsed:
                                contacts.append({
                                    'contact_name': contact_name or '',
                                    'job_title': job_title or '',
                                    'email_address': email,
                                    'email_type': email_type,
                                    'source': f'column_{col}'
                                })
                        else:
                            # 尝试纯邮箱列表解析
                            from utils.file_parser import parse_email_list
                            parsed_list = parse_email_list(text)
                            for email, email_type, contact_name, job_title in parsed_list:
                                contacts.append({
                                    'contact_name': contact_name or '',
                                    'job_title': job_title or '',
                                    'email_address': email,
                                    'email_type': email_type,
                                    'source': f'column_{col}'
                                })

        # 去重
        seen = set()
        unique = []
        for c in contacts:
            email = c['email_address'].lower()
            if email and email not in seen:
                seen.add(email)
                unique.append(c)

        return unique

    # ==================== 4. 数据清洗（优化版）====================

    def clean_results(self, results: List[Dict]) -> List[Dict]:
        """清洗和标准化分析结果"""
        cleaned = []

        for item in results:
            customer = item.get('customer', {})
            contacts = item.get('contacts', [])

            # 清洗客户名称
            customer_name = self._clean_text(customer.get('customer_name', ''))
            if not customer_name or customer_name.lower() in ('nan', 'none', 'null', '', '-'):
                continue

            # 清洗其他字段
            cleaned_customer = {
                'customer_name': customer_name,
                'country': self._clean_text(customer.get('country', '')),
                'website': self._clean_website(customer.get('website', '')),
                'company_info': self._clean_text(customer.get('company_info', '')),
                'address': self._clean_text(customer.get('address', '')),
                'supplier': self._clean_text(customer.get('supplier', '')),
                'customs_data': self._clean_text(customer.get('customs_data', '')),
                'logistics': self._clean_text(customer.get('logistics', ''))
            }

            # 清洗联系人
            cleaned_contacts = []
            for contact in contacts:
                email = contact.get('email_address', '').strip().lower()
                if not email or '@' not in email:
                    continue

                # 验证邮箱格式
                if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    continue

                name = self._clean_text(contact.get('contact_name', ''))
                title = self._clean_text(contact.get('job_title', ''))

                # 检测邮箱类型
                email_type = contact.get('email_type', '')
                if not email_type or email_type not in ('personal', 'public'):
                    if self.is_available():
                        email_type = self.extractor.detect_email_type(email, name)
                    else:
                        email_type = self._rule_based_email_type(email)

                cleaned_contacts.append({
                    'contact_name': name,
                    'job_title': title,
                    'email_address': email,
                    'email_type': email_type,
                    'source': contact.get('source', 'unknown')
                })

            if cleaned_contacts:
                cleaned.append({
                    'customer': cleaned_customer,
                    'contacts': cleaned_contacts,
                    'classification': item.get('classification'),
                    'row_index': item.get('row_index', 0),
                    'sheet_name': item.get('sheet_name', '')
                })

        return cleaned

    def _clean_text(self, text: str) -> str:
        """清洗文本"""
        if not text or pd.isna(text):
            return ''
        text = str(text).strip()
        if text.lower() in ('nan', 'none', 'null', '-', ''):
            return ''
        return text

    def _clean_website(self, url: str) -> str:
        """清洗网址"""
        url = self._clean_text(url)
        if not url:
            return ''
        # 添加 http:// 前缀（如果没有）
        if url and not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        return url

    def _rule_based_email_type(self, email: str) -> str:
        """基于规则的邮箱类型判断（使用新的邮箱分类器）"""
        from utils.email_classifier import classify_email
        email_type, _ = classify_email(email)
        return email_type

    # ==================== 5. 导出结果 ====================

    def export_to_json(self, results: List[Dict], output_path: str):
        """导出结果为 JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"结果已导出到: {output_path}")

    def export_to_excel(self, results: List[Dict], output_path: str):
        """导出结果为 Excel"""
        rows = []
        for item in results:
            customer = item['customer']
            classification = item.get('classification', {})
            for contact in item['contacts']:
                rows.append({
                    '公司名称': customer['customer_name'],
                    '国家': customer.get('country', ''),
                    '网站': customer.get('website', ''),
                    '公司信息': customer.get('company_info', ''),
                    '联系人姓名': contact['contact_name'],
                    '职位': contact['job_title'],
                    '邮箱': contact['email_address'],
                    '邮箱类型': contact['email_type'],
                    '来源': contact['source'],
                    '行业': classification.get('industry', '') if classification else '',
                    '潜力评分': classification.get('potential_score', '') if classification else '',
                    '优先级': classification.get('priority', '') if classification else ''
                })

        df = pd.DataFrame(rows)
        df.to_excel(output_path, index=False)
        print(f"结果已导出到: {output_path}")


# ==================== 便捷函数 ====================

def analyze_excel_file(file_path: str, use_ai: bool = True, sheet_name=None) -> List[Dict]:
    """分析 Excel 文件的便捷函数"""
    analyzer = AIBatchAnalyzer()

    if use_ai and analyzer.is_available():
        print("使用 AI 增强分析...")

    results = analyzer.analyze_file(file_path, sheet_name=sheet_name)
    cleaned = analyzer.clean_results(results)
    print(f"\n分析完成: {len(cleaned)} 个有效客户")

    return cleaned


def quick_analyze(file_path: str) -> Tuple[List[Dict], str]:
    """快速分析文件并返回结果和摘要"""
    try:
        results = analyze_excel_file(file_path)

        total_contacts = sum(len(r['contacts']) for r in results)
        personal_emails = sum(
            1 for r in results for c in r['contacts'] if c['email_type'] == 'personal'
        )
        public_emails = total_contacts - personal_emails

        # 统计 sheet
        sheets = set(r.get('sheet_name', '') for r in results if r.get('sheet_name'))

        summary = f"""
分析完成！
- 客户数量: {len(results)}
- 联系人总数: {total_contacts}
- 个人邮箱: {personal_emails}
- 公共邮箱: {public_emails}
- 涉及 Sheet: {', '.join(sheets) if sheets else '1'}
"""
        return results, summary

    except Exception as e:
        return [], f"分析失败: {str(e)}"


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        results, summary = quick_analyze(file_path)
        print(summary)

        if results:
            output_dir = os.path.dirname(file_path) or '.'
            base_name = os.path.splitext(os.path.basename(file_path))[0]

            json_path = os.path.join(output_dir, f"{base_name}_analysis.json")
            excel_path = os.path.join(output_dir, f"{base_name}_analysis.xlsx")

            analyzer = AIBatchAnalyzer()
            analyzer.export_to_json(results, json_path)
            analyzer.export_to_excel(results, excel_path)
    else:
        print("用法: python ai_batch_analyzer.py <excel_file_path>")
